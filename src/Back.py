import pandas as pd
import openpyxl

class GSL:
    def __init__(self, matrix, j):
        self.method = matrix[0,j]
        self.method_name = matrix[1,j]
        self.time = matrix[3:,j]
        if(matrix[2,j+2]):
            self.sea_level_min = matrix[3:,j+1]
            self.sea_level_mean = matrix[3:,j+2]
            self.sea_level_max = matrix[3:,j+3]
        else:
            self.sea_level_min = matrix[3:,j+1]
            self.sea_level_mean = matrix[3:,j+1]
            self.sea_level_max = matrix[3:,j+1]


class Location:
    #Constructor for a user-entered point
	def __init__(self, name,latitude,longitude,age_min,age_max,min_bathy,max_bathy,altitude):
		self.name = name
		self.latitude = latitude
		self.longitude = longitude
		self.age_min = age_min
		self.age_max = age_max
		self.min_bathy = min_bathy
		self.max_bathy = max_bathy
		self.altitude = altitude

    #Constructor for a point at line i in the matrix
	@classmethod
	def from_file(cls, matrix, i):
		name = matrix[i,2]
		latitude = matrix[i,3]
		longitude = matrix[i,4]
		age_min = matrix[i,7]
		age_max = matrix[i,8]
		min_bathy = matrix[i,10]
		max_bathy = matrix[i,11]
		altitude = matrix[i,12]
		return cls(name,latitude,longitude,age_min,age_max,min_bathy,max_bathy,altitude)

	def __str__(self):
		return ("Location: "+self.name+", "+str(self.latitude)+", "+str(self.longitude)+", "+str(self.age_min)+", "+str(self.age_max)+", "+str(self.min_bathy)+", "+str(self.max_bathy)+", "+str(self.altitude))

    #returns a tab containing all values of vertical movement for this point
	def vertical_movement(self,gsls):
		ret = []
		ret.append(self)
		for gsl in gsls:
			vm = [] #method, method_name, min, mean, max
			time = gsl.time.tolist()
			sl_min = [x for x in gsl.sea_level_min.tolist() if x is not None]
			sl_mean = [x for x in gsl.sea_level_mean.tolist() if x is not None]
			sl_max = [x for x in gsl.sea_level_max.tolist() if x is not None]
			time_filtered = []
			sl_min_filtered = []
			sl_mean_filtered = []
			sl_max_filtered = []
			bool = True
			i = 0

			while(bool):
				if time[i] > self.age_min and time[i] < self.age_max:
					time_filtered.append(time[i])
					sl_min_filtered.append(sl_min[i])
					sl_mean_filtered.append(sl_mean[i])
					sl_max_filtered.append(sl_max[i])
				i+=1
				if(i >= len(time) or time[i] == None):
				    bool = False
				else:
				    if(time[i]>self.age_max):
					bool = False
					
			method = gsl.method
			method_name = gsl.method_name
			vm.append(method)
			vm.append(method_name)

			if len(sl_min_filtered) == 0 or len(sl_mean_filtered) == 0 or len(sl_max_filtered) == 0:
				vm.append(None)
			else:
				min_fvm = self.altitude-max(sl_max_filtered)+self.min_bathy
				mean_fvm = self.altitude-sum(sl_mean_filtered)/float(len(sl_mean_filtered))+((self.min_bathy)+(self.min_bathy))/2
				max_fvm = self.altitude-min(sl_min_filtered)+self.max_bathy
				vm.append(min_fvm)
				vm.append(mean_fvm)
				vm.append(max_fvm)
			ret.append(vm)
		return ret
        #tab contains: Location object as first element and tabs containing [method, method_name, min vm value, mean vm value, max vm value

def read_gsl(file,sheet_name):
    xl = pd.read_excel(file,sheet_name=sheet_name,header=None)
    xl = xl.where((pd.notnull(xl)), None)
    matrix = xl.values
    gsls = []
    for j in range(matrix[0,:].size):
        if matrix[0,j]:
            gsl = GSL(matrix,j)
            gsls.append(gsl)
    return gsls

def read_locations(file,sheet_name):
    xls = pd.ExcelFile(file)
    xl = pd.read_excel(xls,sheet_name=sheet_name,header=None)
    xl = xl.where((pd.notnull(xl)), None)
    matrix = xl.values
    locations = []
    for i in range(4,matrix[:,1].size):
        if matrix[i,1]:
            loc = Location.from_file(matrix,i)
            locations.append(loc)
    return locations


def export(filename,sheetname,fvms):
    i = 5

    wb = load_workbook(filename)
    ws = wb[sheetname]
    for fvm in fvms:
        col = 18
        while(ws.cell(row=i, column=3).value is not None):
            i+=1
        ws.cell(row=i,column=3).value = fvm[0].name
        ws.cell(row=i,column=4).value = fvm[0].latitude
        ws.cell(row=i,column=5).value = fvm[0].longitude
        ws.cell(row=i,column=8).value = fvm[0].age_min
        ws.cell(row=i,column=9).value = fvm[0].age_max
        ws.cell(row=i,column=11).value = fvm[0].min_bathy
        ws.cell(row=i,column=12).value = fvm[0].max_bathy
        ws.cell(row=i,column=13).value = fvm[0].altitude
        for j in range (1,len(fvm)):
            ws.cell(row=1,column=col-3).value = fvm[j][0]
            ws.cell(row=2,column=col-3).value = fvm[j][1]
            if(fvm[j][2] is not None):
                ws.cell(row=i,column=col).value = fvm[j][2]
                ws.cell(row=i,column=col+1).value = fvm[j][3]
                ws.cell(row=i,column=col+2).value = fvm[j][4]
            col+=7
    wb.save(filename)
    wb.close()
